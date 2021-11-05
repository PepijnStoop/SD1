# the goal of this program is to add and delete data to an database (Excel file) from a source Excel file
# to create the start of a inventory management program.

# imports openpyxl to make Python work with Excel
# imports classes and mapping to make sure that there are no "magic numbers"
import openpyxl as xl
from classes import Product
from mapping import PRODUCT_SNR, PRODUCT_TYPE, PRODUCT_SOORT, PRODUCT_MAGAZIJN, PRODUCT_ID


# load main Excel file
mainExcel = "magazijn.xlsx"
workbookMain = xl.load_workbook(mainExcel, read_only=False)
sheetMain = workbookMain.worksheets[0]

# without deleting the extra empty columns the print function will print 15 by 100 columns/rows
sheetMain.delete_cols(idx=6, amount=10)
maxRows = sheetMain.max_row
maxCols = sheetMain.max_column

# the code below is used in a multiple functions to detect if data is already in the database or not
# so made it a global variable to save code
existingRows = []
for column in sheetMain['C']:
    existingRows.append(column.value)

# easy function to print the main Excel file
def printRows():
    for row in sheetMain.iter_rows(values_only=True):
        print(row)


# copying everything from the source Excel to the main Excel without overwriting the existing data
def addSource():
    sourceExcel = "source.xlsx"
    workbookSource = xl.load_workbook(sourceExcel, read_only=False)
    sheetSource = workbookSource.worksheets[0]
  
    for currentRow in range(1,maxRows+1):
        cellValue = sheetSource.cell(row=currentRow+1, column=3)
        if cellValue.value in existingRows:
            print("Duplicate data detected!: " +str(cellValue.value))
            print("List not updated!")
            break
        else:
            for i in range(2, maxRows + 1):
                for j in range(1, maxCols + 1):
                    copy = sheetSource.cell(row=i, column=j)
                    sheetMain.cell(row=maxRows + i, column=j).value = copy.value
    # sourceExcel = "source.xlsx"
    # workbookSource = xl.load_workbook(sourceExcel, read_only=False)
    # sheetSource = workbookSource.worksheets[0]
  
    # existingRows = []
    # for column in sheetMain['C']:
    #     existingRows.append(column.value)

    # for currentRow in range(1,maxRows+1):
    #     cellValue = sheetSource.cell(row=currentRow+1, column=3)
    #     if cellValue.value in existingRows:
    #         print("Duplicate data detected!: " +str(cellValue.value))

    # for i in range(2, maxRows + 1):
    #     for j in range(1, maxCols + 1):
    #         copy = sheetSource.cell(row=i, column=j)
    #         sheetMain.cell(row=maxRows + i, column=j).value = copy.value
    

# deleting empty rows from sheetMain
def deleteEmpty():
    indexRow = []

    # loop each row in column A
    for i in range(1, sheetMain.max_row + 1):
        # define emptiness of cell
        if sheetMain.cell(i, 3).value is None:
            # collect indexes of rows
            indexRow.append(i)

    # loop each index value
    for rowDel in range(len(indexRow)):
        sheetMain.delete_rows(idx=indexRow[rowDel], amount=1)
        # exclude offset of rows through each iteration
        indexRow = list(map(lambda k: k - 1, indexRow))

def removeInventory():
    removeExcel = "remove.xlsx"
    workbookRemove = xl.load_workbook(removeExcel, read_only=False)
    sheetRemove = workbookRemove.worksheets[0]
    sheetRemove.delete_rows(idx=1, amount=1)

    for currentRow in range(1,maxRows+1):
        cellValue = sheetRemove.cell(row=currentRow+1, column=3)
        if cellValue.value not in existingRows:
            print("Can't remove asset as it does not exist in the list!: " +str(cellValue.value))
            print("List not updated!")
            break
        else:
            removeRow = []
            for column in sheetRemove['C']:
                removeRow.append(column.value)


            for currentRow in range(1,maxRows+1):
                cellValue = sheetMain.cell(row=currentRow, column=3)
                if cellValue.value in removeRow:
                    cellValue.value = None
    


# list with the products
def printProducts():
    products = []
    for row in sheetMain.iter_rows(min_row=1, values_only=True):
        product = Product(magazijn=row[PRODUCT_MAGAZIJN],
                          soort=row[PRODUCT_SOORT],
                          serienummer=row[PRODUCT_SNR],
                          type=row[PRODUCT_TYPE],
                          id=row[PRODUCT_ID])
        products.append(product)

    printRows()



# main function to run the program
def main():
    # defining the variable "option" to 0 so it can be used in the while loop
    option = 0
    while option != 4:
        print("What do you want to do?")
        print("1. List current inventory.")
        print("2. Add to current inventory.")
        print("3. Delete current inventory.")
        print("4. Exit program.")
        option = int(input("Select option: "))

        if option == 1:
            deleteEmpty()
            printProducts()

        elif option == 2:
            addSource()
            deleteEmpty()
#            printProducts()
            workbookMain.save('test.xlsx')

        elif option == 3:
            removeInventory()
            deleteEmpty()
            workbookMain.save('test2.xlsx')
#            printProducts()
#            print("This feature is currently unavailable.")

        elif option == 4:
            exit(0)
        else:
            while option < 1 or option > 4:
                option = int(input("Please select either option 1, 2, 3 or 4"))

# run the program
main()
